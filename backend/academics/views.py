from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from django.utils import timezone
import csv
from io import StringIO
from io import BytesIO
from datetime import date
from django.http import HttpResponse
import re

from .models import (
    Department, Program, Semester, Course, Batch, Section, StudentGroup,
    SubjectDomain, FacultyProfile, FacultyAvailability, FacultyPreference,
    FacultyPreferredSubject, FacultyEligibleSubject, FacultySubjectAssignment,
    CourseSection, DepartmentRoomPreference
)
from .serializers import (
    DepartmentSerializer, SubjectDomainSerializer, ProgramSerializer, SemesterSerializer, CourseSerializer,
    BatchSerializer, SectionSerializer, StudentGroupSerializer,
    FacultyProfileSerializer, FacultyProfessionalDetailSerializer,
    FacultyAvailabilitySerializer, FacultyPreferenceSerializer,
    FacultyEligibleSubjectSerializer, FacultySubjectAssignmentSerializer,
    CourseSectionSerializer, DepartmentRoomPreferenceSerializer
)
from .syllabus_parser import parse_syllabus_pages
from .matching import profile_completion_state, generate_faculty_eligible_subjects, regenerate_department_eligibility
from accounts.models import CustomUser, Role


# ── Base ─────────────────────────────────────────────────────────────────────

class TenantAcademicViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    pagination_class = None

    def get_tenant(self):
        return self.request.user.tenant

    def _is_admin(self):
        role = getattr(getattr(self.request.user, 'role', None), 'name', '')
        return role in ['super_admin', 'tenant_admin', 'academic_admin']

    def _role_name(self):
        return getattr(getattr(self.request.user, 'role', None), 'name', '')

    def _get_hod_department(self):
        try:
            if self.request.user.faculty_profile.department_id:
                return self.request.user.faculty_profile.department
        except Exception:
            pass

        department = Department.objects.filter(
            tenant=self.get_tenant(),
            head_of_department=self.request.user,
        ).select_related('campus').first()
        if department:
            return department

        role_name = self._role_name()
        if role_name != 'hod':
            return None

        normalized_user_department = self._normalize_department_name(getattr(self.request.user, 'department', ''))
        if not normalized_user_department:
            return None

        department_qs = Department.objects.filter(tenant=self.get_tenant()).select_related('campus')
        if getattr(self.request.user, 'campus_id', None):
            campus_matches = department_qs.filter(campus_id=self.request.user.campus_id)
            for item in campus_matches:
                if self._normalize_department_name(item.name) == normalized_user_department:
                    return item

        for item in department_qs:
            if self._normalize_department_name(item.name) == normalized_user_department:
                return item

        return None

    def _normalize_department_name(self, value):
        if not value:
            return ''
        base = value.split('(')[0].strip().lower().replace('&', 'and')
        return ' '.join(base.split())

    def _get_hod_department_scope_ids(self):
        department = self._get_hod_department()
        if not department:
            return []
        normalized_name = self._normalize_department_name(department.name)
        departments = Department.objects.filter(tenant=self.get_tenant())
        if department.campus_id:
            departments = departments.filter(campus_id=department.campus_id)
        matches = []
        for item in departments:
            if self._normalize_department_name(item.name) == normalized_name:
                matches.append(item.id)
        return matches or [department.id]


# ── Department ────────────────────────────────────────────────────────────────

class DepartmentViewSet(TenantAcademicViewSet):
    serializer_class = DepartmentSerializer
    filterset_fields = ['id', 'campus', 'campus_id', 'is_active']
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'code']

    def get_queryset(self):
        qs = Department.objects.filter(tenant=self.get_tenant()).select_related('head_of_department')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(id=hod_department.id)
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())


# ── Program ───────────────────────────────────────────────────────────────────

class SubjectDomainViewSet(TenantAcademicViewSet):
    serializer_class = SubjectDomainSerializer
    filterset_fields = ['department', 'department_id', 'is_active']
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'code']

    def get_queryset(self):
        if not self._get_hod_department() and not self._is_admin():
            return SubjectDomain.objects.none()
        qs = SubjectDomain.objects.filter(tenant=self.get_tenant()).select_related('department__campus')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(department_id__in=self._get_hod_department_scope_ids())
        program_id = self.request.query_params.get('program_id')
        if program_id:
            qs = qs.filter(
                Q(primary_courses__semester__program_id=program_id) |
                Q(secondary_courses__semester__program_id=program_id)
            ).distinct()
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())


class ProgramViewSet(TenantAcademicViewSet):
    serializer_class = ProgramSerializer
    filterset_fields = ['department', 'department_id', 'department__campus', 'department__campus_id', 'degree_type', 'is_active']
    search_fields = ['name', 'code']
    ordering_fields = ['name']

    def get_queryset(self):
        qs = Program.objects.filter(tenant=self.get_tenant()).select_related('department')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(department_id__in=self._get_hod_department_scope_ids())
        return qs

    def _is_hod_scoped_user(self):
        return bool(self._get_hod_department()) and not self._is_admin()

    def _extract_pdf_text(self, file_path):
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise RuntimeError('PDF parser dependency is not available.') from exc

        reader = PdfReader(file_path)
        chunks = []
        for page in reader.pages:
            chunks.append(page.extract_text() or '')
        return '\n'.join(chunks)

    def _extract_pdf_pages(self, file_path):
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise RuntimeError('PDF parser dependency is not available.') from exc

        reader = PdfReader(file_path)
        return [page.extract_text() or '' for page in reader.pages]

    def _current_academic_year(self):
        year = date.today().year
        return f'{year}-{str(year + 1)[-2:]}'

    def _build_syllabus_template_workbook(self, program):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        workbook = Workbook()
        workbook.remove(workbook.active)

        headers = [
            'Subject Code',
            'Subject Name',
            'Subject Domain',
            'Optional Secondary Domain',
            'Theory Hours (L)',
            'Tutorial Hours (T)',
            'Lab Hours (P)',
            'Credits',
            'Course Type',
            'Is Elective',
            'Description',
        ]

        for semester_number in range(1, (program.total_semesters or 0) + 1):
            sheet = workbook.create_sheet(title=f'Semester {semester_number}')
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1F3A5F')
            sheet.freeze_panes = 'A2'
            sheet.append(['', '', '', '', 0, 0, 0, 0, 'theory', 'No', ''])

            widths = {
                'A': 20, 'B': 40, 'C': 24, 'D': 24, 'E': 18,
                'F': 18, 'G': 18, 'H': 12, 'I': 18, 'J': 14, 'K': 32,
            }
            for column, width in widths.items():
                sheet.column_dimensions[column].width = width

        return workbook

    def _parse_syllabus_workbook(self, file_path):
        from openpyxl import load_workbook

        workbook = load_workbook(filename=file_path, data_only=True)
        extracted_subjects = []

        expected_headers = {
            'subject code': 'subject_code',
            'subject name': 'subject_name',
            'subject domain': 'primary_domain',
            'optional secondary domain': 'secondary_domain',
            'theory hours (l)': 'lecture_hours',
            'tutorial hours (t)': 'tutorial_hours',
            'lab hours (p)': 'practical_hours',
            'credits': 'credits',
            'course type': 'course_type',
            'is elective': 'is_elective',
            'description': 'description',
        }
        errors = []

        for sheet in workbook.worksheets:
            match = re.search(r'(\d+)', sheet.title or '')
            if not match:
                continue

            semester_number = int(match.group(1))
            header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            header_map = {}
            for idx, header in enumerate(header_row):
                normalized = str(header or '').strip().lower()
                if normalized in expected_headers:
                    header_map[expected_headers[normalized]] = idx

            if 'subject_code' not in header_map or 'subject_name' not in header_map or 'primary_domain' not in header_map:
                continue

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                subject_code = str(row[header_map['subject_code']] or '').strip()
                subject_name = str(row[header_map['subject_name']] or '').strip()
                primary_domain = str(row[header_map['primary_domain']] or '').strip()
                if not subject_code and not subject_name:
                    continue
                if not subject_code or not subject_name:
                    continue
                if not primary_domain:
                    errors.append(f'{sheet.title} row {row_number}: Subject Domain is required for {subject_code or subject_name}.')
                    continue

                def _num(key):
                    idx = header_map.get(key)
                    value = row[idx] if idx is not None and idx < len(row) else 0
                    try:
                        return int(float(value or 0))
                    except Exception:
                        return 0

                credits_idx = header_map.get('credits')
                credits_value = row[credits_idx] if credits_idx is not None and credits_idx < len(row) else 0
                try:
                    credits = float(credits_value or 0)
                except Exception:
                    credits = 0

                extracted_subjects.append({
                    'semester_number': semester_number,
                    'subject_code': subject_code.upper(),
                    'subject_name': subject_name,
                    'primary_domain': primary_domain,
                    'secondary_domain': str(row[header_map['secondary_domain']] or '').strip() if 'secondary_domain' in header_map else '',
                    'lecture_hours': _num('lecture_hours'),
                    'tutorial_hours': _num('tutorial_hours'),
                    'practical_hours': _num('practical_hours'),
                    'credits': credits,
                    'course_type': str(row[header_map['course_type']] or '').strip().lower() if 'course_type' in header_map else '',
                    'is_elective': str(row[header_map['is_elective']] or '').strip().lower() in ['yes', 'true', '1'] if 'is_elective' in header_map else False,
                    'description': str(row[header_map['description']] or '').strip() if 'description' in header_map else '',
                })

        if errors:
            raise ValueError(' | '.join(errors[:10]))
        return extracted_subjects

    def _normalize_domain_name(self, value):
        return ' '.join(str(value or '').strip().split())

    def _resolve_domain_for_program(self, program, domain_value, allow_blank=False, create_missing=False):
        cleaned = self._normalize_domain_name(domain_value)
        if not cleaned:
            return None if allow_blank else None
        domain = SubjectDomain.objects.filter(
            tenant=program.tenant,
            department=program.department,
        ).filter(
            Q(name__iexact=cleaned) | Q(code__iexact=cleaned)
        ).first()
        if domain:
            if not domain.is_active:
                domain.is_active = True
                domain.save(update_fields=['is_active'])
            return domain
        if not create_missing:
            return None
        return SubjectDomain.objects.create(
            tenant=program.tenant,
            department=program.department,
            name=cleaned,
            description=f'Auto-created from syllabus import for {program.name}.',
            is_active=True,
        )

    def _ensure_import_domains(self, program, extracted_subjects):
        for row in extracted_subjects:
            self._resolve_domain_for_program(program, row.get('primary_domain'), create_missing=True)
            self._resolve_domain_for_program(
                program,
                row.get('secondary_domain'),
                allow_blank=True,
                create_missing=True,
            )

    def _sync_program_subjects(self, program, extracted_subjects):
        academic_year = self._current_academic_year()
        semester_cache = {}
        validation_errors = []
        self._ensure_import_domains(program, extracted_subjects)

        for row in extracted_subjects:
            semester_number = row.get('semester_number')
            if semester_number:
                semester = semester_cache.get(semester_number)
                if semester is None:
                    semester, _ = Semester.objects.get_or_create(
                        tenant=program.tenant,
                        program=program,
                        semester_number=semester_number,
                        academic_year=academic_year,
                        defaults={
                            'name': f'Semester {semester_number}',
                            'term': 'odd' if semester_number % 2 else 'even',
                        }
                    )
                    semester_cache[semester_number] = semester
            else:
                semester = None

            lecture_hours = row.get('lecture_hours') or 0
            tutorial_hours = row.get('tutorial_hours') or 0
            practical_hours = row.get('practical_hours') or 0

            if practical_hours and lecture_hours == 0 and tutorial_hours == 0:
                course_type = 'practical'
            elif tutorial_hours and not practical_hours:
                course_type = 'tutorial'
            else:
                course_type = 'theory'

            primary_domain = self._resolve_domain_for_program(
                program,
                row.get('primary_domain'),
                create_missing=True,
            )
            secondary_domain = self._resolve_domain_for_program(
                program,
                row.get('secondary_domain'),
                allow_blank=True,
                create_missing=True,
            )
            if not primary_domain:
                validation_errors.append(f"Domain '{row.get('primary_domain')}' is not defined for {program.department.name}.")
                continue

            Course.objects.update_or_create(
                tenant=program.tenant,
                code=row['subject_code'],
                defaults={
                    'department': program.department,
                    'semester': semester,
                    'primary_domain': primary_domain,
                    'secondary_domain': secondary_domain,
                    'name': row['subject_name'],
                    'course_type': course_type,
                    'credits': row.get('credits') or 0,
                    'lecture_hours': lecture_hours,
                    'tutorial_hours': tutorial_hours,
                    'practical_hours': practical_hours,
                    'is_elective': bool(row.get('is_elective', False)),
                    'is_active': True,
                      'description': row.get('description') or f'Imported from syllabus for {program.name}',
                  }
              )
        if validation_errors:
            raise ValueError(' | '.join(validation_errors[:10]))
        regenerate_department_eligibility(program.tenant, [program.department_id])

    def _syllabus_import_stats(self, program):
        parsed_rows = program.syllabus_extracted_subjects or []
        parsed_count = len(parsed_rows)
        stored_count = Course.objects.filter(tenant=program.tenant, semester__program=program).count()
        unique_codes_count = len({
            (row.get('subject_code') or '').strip().upper()
            for row in parsed_rows
            if (row.get('subject_code') or '').strip()
        })
        duplicate_rows_count = max(parsed_count - unique_codes_count, 0)
        return {
            'parsed_rows_count': parsed_count,
            'stored_subjects_count': stored_count,
            'duplicate_code_rows_count': duplicate_rows_count,
        }

    def _reparse_and_sync_program(self, program):
        file_name = (program.syllabus_file.name or '').lower()
        if file_name.endswith('.xlsx') or file_name.endswith('.xlsm') or file_name.endswith('.xltx') or file_name.endswith('.xltm'):
            extracted_subjects = self._parse_syllabus_workbook(program.syllabus_file.path)
            raw_text = f'Workbook import for {program.name}'
        else:
            page_texts = self._extract_pdf_pages(program.syllabus_file.path)
            raw_text = '\n'.join(page_texts)
            extracted_subjects = parse_syllabus_pages(page_texts)
        program.syllabus_raw_text = raw_text
        program.syllabus_extracted_subjects = extracted_subjects
        program.syllabus_parse_status = 'parsed' if extracted_subjects else 'failed'
        program.syllabus_last_error = '' if extracted_subjects else 'No subject rows could be extracted from this syllabus format.'
        program.syllabus_updated_at = timezone.now()
        program.save(update_fields=[
            'syllabus_raw_text', 'syllabus_extracted_subjects',
            'syllabus_parse_status', 'syllabus_last_error', 'syllabus_updated_at'
        ])
        if extracted_subjects:
            self._sync_program_subjects(program, extracted_subjects)
        return extracted_subjects

    @action(detail=False, methods=['POST'], url_path='bulk-delete')
    def bulk_delete(self, request):
        if not self._is_admin() and not self._get_hod_department():
            return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get('ids', [])
        if not isinstance(ids, list) or not ids:
            return Response({'error': 'ids must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

        qs = self.get_queryset().filter(id__in=ids)
        deleted = 0
        failed = []
        for program in qs:
            try:
                program.delete()
                deleted += 1
            except Exception as exc:
                failed.append({'id': str(program.id), 'name': program.name, 'error': str(exc)})

        return Response({'deleted_count': deleted, 'failed': failed})

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())

    def create(self, request, *args, **kwargs):
        if self._is_hod_scoped_user():
            return Response({'detail': 'HOD can manage syllabus only for department programs.'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if self._is_hod_scoped_user():
            return Response({'detail': 'HOD can manage syllabus only for department programs.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if self._is_hod_scoped_user():
            allowed_fields = {'syllabus_overview', 'syllabus_document_url'}
            incoming_fields = set(request.data.keys())
            disallowed_fields = incoming_fields - allowed_fields
            if disallowed_fields:
                return Response(
                    {'detail': 'HOD can update only syllabus fields.', 'fields': sorted(disallowed_fields)},
                    status=status.HTTP_403_FORBIDDEN
                )
            serializer = self.get_serializer(self.get_object(), data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save(syllabus_updated_at=timezone.now())
            return Response(serializer.data)
        return super().partial_update(request, *args, **kwargs)

    @action(detail=True, methods=['POST'], url_path='upload-syllabus')
    def upload_syllabus(self, request, pk=None):
        program = self.get_object()
        if not self._is_hod_scoped_user() and not self._is_admin():
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        syllabus_file = request.FILES.get('file')
        if not syllabus_file:
            return Response({'detail': 'Syllabus file is required.'}, status=status.HTTP_400_BAD_REQUEST)

        program.syllabus_file = syllabus_file
        if 'syllabus_overview' in request.data:
            program.syllabus_overview = request.data.get('syllabus_overview', '')
        if 'syllabus_document_url' in request.data:
            program.syllabus_document_url = request.data.get('syllabus_document_url', '')
        program.syllabus_parse_status = 'uploaded'
        program.syllabus_last_error = ''
        program.syllabus_raw_text = ''
        program.syllabus_extracted_subjects = []
        program.syllabus_updated_at = timezone.now()
        program.save()

        try:
            self._reparse_and_sync_program(program)
        except Exception as exc:
            program.syllabus_parse_status = 'failed'
            program.syllabus_last_error = str(exc)
            program.save(update_fields=['syllabus_parse_status', 'syllabus_last_error'])
            serializer = self.get_serializer(program)
            return Response({
                'detail': str(exc),
                **serializer.data,
                **self._syllabus_import_stats(program),
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(program)
        return Response({
            **serializer.data,
            **self._syllabus_import_stats(program),
        })

    @action(detail=True, methods=['POST'], url_path='sync-syllabus-subjects')
    def sync_syllabus_subjects(self, request, pk=None):
        program = self.get_object()
        if not self._is_hod_scoped_user() and not self._is_admin():
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        if not program.syllabus_file:
            return Response({'detail': 'No uploaded syllabus found for this programme.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            self._reparse_and_sync_program(program)
        except Exception as exc:
            program.syllabus_parse_status = 'failed'
            program.syllabus_last_error = str(exc)
            program.save(update_fields=['syllabus_parse_status', 'syllabus_last_error'])
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(program)
        return Response({
            **serializer.data,
            **self._syllabus_import_stats(program),
        })

    @action(detail=True, methods=['POST'], url_path='delete-syllabus')
    def delete_syllabus(self, request, pk=None):
        program = self.get_object()
        if not self._is_hod_scoped_user() and not self._is_admin():
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        if program.syllabus_file:
            try:
                program.syllabus_file.delete(save=False)
            except Exception:
                pass

        deleted_courses = 0
        deleted_semesters = 0
        course_qs = Course.objects.filter(tenant=program.tenant, semester__program=program)
        deleted_courses = course_qs.count()
        course_qs.delete()

        semester_qs = Semester.objects.filter(tenant=program.tenant, program=program)
        deleted_semesters = semester_qs.count()
        semester_qs.delete()

        program.syllabus_file = None
        program.syllabus_overview = ''
        program.syllabus_document_url = ''
        program.syllabus_parse_status = 'not_uploaded'
        program.syllabus_raw_text = ''
        program.syllabus_extracted_subjects = []
        program.syllabus_last_error = ''
        program.syllabus_updated_at = timezone.now()
        program.save(update_fields=[
            'syllabus_file', 'syllabus_overview', 'syllabus_document_url',
            'syllabus_parse_status', 'syllabus_raw_text',
            'syllabus_extracted_subjects', 'syllabus_last_error',
            'syllabus_updated_at'
        ])

        serializer = self.get_serializer(program)
        return Response({
            **serializer.data,
            'deleted_courses_count': deleted_courses,
            'deleted_semesters_count': deleted_semesters,
        })

    @action(detail=True, methods=['GET'], url_path='download-syllabus-template')
    def download_syllabus_template(self, request, pk=None):
        program = self.get_object()
        if not self._is_hod_scoped_user() and not self._is_admin():
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        workbook = self._build_syllabus_template_workbook(program)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        safe_name = ''.join(ch if ch.isalnum() else '_' for ch in program.name).strip('_') or 'programme'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_name}_syllabus_template.xlsx"'
        return response


# ── Semester ──────────────────────────────────────────────────────────────────

class SemesterViewSet(TenantAcademicViewSet):
    serializer_class = SemesterSerializer
    filterset_fields = ['program', 'program_id', 'program__department', 'program__department_id', 'program__department__campus', 'program__department__campus_id', 'academic_year', 'is_current', 'term']
    search_fields = ['name', 'academic_year']
    ordering_fields = ['semester_number']

    def get_queryset(self):
        qs = Semester.objects.filter(tenant=self.get_tenant()).select_related('program')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(program__department_id__in=self._get_hod_department_scope_ids())
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())

    @action(detail=True, methods=['POST'], url_path='set-current')
    def set_current(self, request, pk=None):
        """Mark a semester as current and unmark others in the same program."""
        sem = self.get_object()
        Semester.objects.filter(tenant=self.get_tenant(), program=sem.program, is_current=True).update(is_current=False)
        sem.is_current = True
        sem.save()
        return Response({'status': f'{sem.name} is now the current semester.'})


# ── Course ────────────────────────────────────────────────────────────────────

class CourseViewSet(TenantAcademicViewSet):
    serializer_class = CourseSerializer
    filterset_fields = ['department', 'department_id', 'department__campus', 'department__campus_id', 'semester', 'semester_id', 'course_type', 'is_elective', 'is_active', 'primary_domain', 'secondary_domain']
    search_fields = ['name', 'code']
    ordering_fields = ['code', 'name']

    def get_queryset(self):
        qs = Course.objects.filter(tenant=self.get_tenant()).select_related('department', 'semester__program', 'primary_domain', 'secondary_domain')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(department_id__in=self._get_hod_department_scope_ids())
        program_id = self.request.query_params.get('program_id')
        semester_number = self.request.query_params.get('semester_number')
        domain_id = self.request.query_params.get('domain_id')
        if program_id:
            qs = qs.filter(semester__program_id=program_id)
        if semester_number:
            qs = qs.filter(semester__semester_number=semester_number)
        if domain_id:
            qs = qs.filter(Q(primary_domain_id=domain_id) | Q(secondary_domain_id=domain_id))
        return qs

    @action(detail=False, methods=['POST'], url_path='bulk-delete')
    def bulk_delete(self, request):
        if not self._is_admin() and not self._get_hod_department():
            return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get('ids', [])
        if not isinstance(ids, list) or not ids:
            return Response({'error': 'ids must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)

        qs = self.get_queryset().filter(id__in=ids)
        deleted = 0
        failed = []
        for course in qs:
            try:
                course.delete()
                deleted += 1
            except Exception as exc:
                failed.append({'id': str(course.id), 'name': course.name, 'error': str(exc)})

        return Response({'deleted_count': deleted, 'failed': failed})

    def perform_create(self, serializer):
        course = serializer.save(tenant=self.get_tenant())
        regenerate_department_eligibility(course.tenant, [course.department_id])

    def perform_update(self, serializer):
        course = serializer.save()
        regenerate_department_eligibility(course.tenant, [course.department_id])

    def perform_destroy(self, instance):
        tenant = instance.tenant
        department_id = instance.department_id
        super().perform_destroy(instance)
        regenerate_department_eligibility(tenant, [department_id])

    @action(detail=False, methods=['POST'], url_path='bulk-upload')
    def bulk_upload(self, request):
        """CSV bulk import for courses.
        Required columns: code, name, course_type, credits, department_code
        Optional: semester_number, lecture_hours, tutorial_hours, practical_hours, is_elective
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        tenant = self.get_tenant()
        csv_data = StringIO(file.read().decode('utf-8'))
        reader = csv.DictReader(csv_data)
        created, skipped, errors = 0, 0, []

        for row_num, row in enumerate(reader, start=2):
            code = row.get('code', '').strip().upper()
            name = row.get('name', '').strip()
            course_type = row.get('course_type', 'theory').strip().lower()
            credits = row.get('credits', '3').strip()
            dept_code = row.get('department_code', '').strip().upper()

            if not all([code, name, dept_code]):
                errors.append(f'Row {row_num}: Missing code, name, or department_code.')
                continue

            try:
                dept = Department.objects.get(tenant=tenant, code=dept_code)
            except Department.DoesNotExist:
                errors.append(f'Row {row_num}: Department "{dept_code}" not found.')
                continue

            if Course.objects.filter(tenant=tenant, code=code).exists():
                skipped += 1
                errors.append(f'Row {row_num}: Course {code} already exists (skipped).')
                continue

            try:
                Course.objects.create(
                    tenant=tenant,
                    department=dept,
                    code=code,
                    name=name,
                    course_type=course_type,
                    credits=float(credits),
                    lecture_hours=int(row.get('lecture_hours', 3)),
                    tutorial_hours=int(row.get('tutorial_hours', 1)),
                    practical_hours=int(row.get('practical_hours', 0)),
                    is_elective=row.get('is_elective', 'false').strip().lower() == 'true',
                )
                created += 1
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')

        return Response({'message': f'Done: {created} created, {skipped} skipped.', 'errors': errors})


# ── Batch & Section ───────────────────────────────────────────────────────────

class BatchViewSet(TenantAcademicViewSet):
    serializer_class = BatchSerializer
    filterset_fields = ['program', 'program_id', 'program__department', 'program__department_id', 'program__department__campus', 'program__department__campus_id', 'is_active']
    search_fields = ['name']
    ordering_fields = ['start_year']

    def get_queryset(self):
        qs = Batch.objects.filter(tenant=self.get_tenant()).select_related('program')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(program__department_id__in=self._get_hod_department_scope_ids())
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())


class SectionViewSet(TenantAcademicViewSet):
    serializer_class = SectionSerializer
    filterset_fields = ['batch', 'batch__program', 'batch__program__department', 'batch__program__department__campus_id', 'batch__program__department_id', 'is_active']
    search_fields = ['name']

    def get_queryset(self):
        qs = Section.objects.filter(tenant=self.get_tenant()).select_related('batch', 'class_advisor')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(batch__program__department_id__in=self._get_hod_department_scope_ids())
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())


class StudentGroupViewSet(TenantAcademicViewSet):
    serializer_class = StudentGroupSerializer
    filterset_fields = ['section']

    def get_queryset(self):
        return StudentGroup.objects.filter(tenant=self.get_tenant()).select_related('section')

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())


# ── Faculty ───────────────────────────────────────────────────────────────────

class FacultyProfileViewSet(TenantAcademicViewSet):
    serializer_class = FacultyProfileSerializer
    filterset_fields = ['department', 'department_id', 'department__campus', 'department__campus_id', 'designation', 'status']
    search_fields = ['user__full_name', 'user__email', 'employee_id', 'specialization']
    ordering_fields = ['user__full_name', 'designation']

    def get_queryset(self):
        qs = FacultyProfile.objects.filter(
            tenant=self.get_tenant(),
        ).select_related('user', 'department__campus')
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            qs = qs.filter(department=hod_department)

        campus_id = self.request.query_params.get('campus_id') or self.request.query_params.get('department__campus_id')
        department_id = self.request.query_params.get('department_id') or self.request.query_params.get('department')
        designation = self.request.query_params.get('designation')

        if campus_id:
            qs = qs.filter(department__campus_id=campus_id)
        if department_id:
            qs = qs.filter(department_id=department_id)
        if designation:
            qs = qs.filter(designation=designation)
        return qs

    def _normalize_department_name(self, value):
        return super()._normalize_department_name(value)

    def _resolve_department_for_user(self, user, departments_by_campus, all_departments):
        try:
            profile = user.faculty_profile
            if profile.department_id:
                return profile.department
        except FacultyProfile.DoesNotExist:
            profile = None

        normalized_user_department = self._normalize_department_name(user.department)
        if not normalized_user_department:
            return None

        # 1. First, try to match within the user's explicit campus
        if user.campus_id:
            campus_departments = departments_by_campus.get(str(user.campus_id), [])
            for department in campus_departments:
                if self._normalize_department_name(department.name) == normalized_user_department:
                    return department

        # 2. If no match in their explicit campus, or no campus on user, search all departments
        # BUT prioritize departments that belong to ANY campus rather than no campus,
        # or exactly match if it's the only one.
        matches = []
        for department in all_departments:
            if self._normalize_department_name(department.name) == normalized_user_department:
                matches.append(department)

        if matches:
            # If multiple matches (e.g., 'civil engineering' in Main and North campus),
            # but user has no campus_id, we just have to pick the first one (or ideally we shouldn't have users without campus_id).
            return matches[0]

        return None

    def _serialize_directory_user(self, user, department):
        try:
            profile = user.faculty_profile
        except FacultyProfile.DoesNotExist:
            profile = None

        campus = user.campus or getattr(department, 'campus', None)
        status = profile.status if profile else ('active' if user.is_active else 'inactive')
        status_display = profile.get_status_display() if profile else ('Active' if user.is_active else 'Inactive')

        return {
            'id': user.id,
            'profile_id': str(profile.id) if profile else None,
            'user': user.id,
            'user_name': user.full_name,
            'user_email': user.email,
            'employee_id': profile.employee_id if profile else '',
            'campus_id': str(campus.id) if campus else None,
            'campus_name': campus.name if campus else None,
            'department': str(department.id) if department else None,
            'department_name': department.name if department else (user.department or ''),
            'designation': profile.designation if profile else '',
            'designation_display': profile.get_designation_display() if profile else '',
            'specialization': profile.specialization if profile else '',
            'primary_specialization_domain': str(profile.primary_specialization_domain_id) if profile and profile.primary_specialization_domain_id else None,
            'primary_specialization_domain_name': profile.primary_specialization_domain.name if profile and profile.primary_specialization_domain_id else '',
            'secondary_specialization_domain_names': [item.name for item in profile.secondary_specialization_domains.all()] if profile else [],
            'qualifications': profile.qualifications if profile else '',
            'skills': profile.skills if profile else '',
            'years_of_experience': profile.years_of_experience if profile else None,
            'max_weekly_hours': profile.max_weekly_hours if profile else None,
            'status': status,
            'status_display': status_display,
            'profile_completion': profile_completion_state(profile)['percentage'] if profile else 0,
            'profile_is_complete': profile_completion_state(profile)['ready_for_preferences'] if profile else False,
        }

    def list(self, request, *args, **kwargs):
        tenant = self.get_tenant()
        campus_id = request.query_params.get('campus_id') or request.query_params.get('department__campus_id')
        department_id = request.query_params.get('department_id') or request.query_params.get('department')
        designation = request.query_params.get('designation')
        search = (request.query_params.get('search') or '').strip().lower()

        users_qs = CustomUser.objects.filter(
            tenant=tenant,
            role__name='faculty',
        ).select_related('campus', 'role').order_by('full_name')

        if campus_id:
            users_qs = users_qs.filter(campus_id=campus_id)

        departments = list(
            Department.objects.filter(tenant=tenant).select_related('campus').order_by('name')
        )
        departments_by_campus = {}
        for department in departments:
            key = str(department.campus_id) if department.campus_id else None
            departments_by_campus.setdefault(key, []).append(department)

        rows = []
        for user in users_qs:
            department = self._resolve_department_for_user(user, departments_by_campus, departments)
            row = self._serialize_directory_user(user, department)

            if department_id and str(row['department']) != str(department_id):
                continue
            if designation and row['designation'] != designation:
                continue
            if search:
                haystack = ' '.join([
                    row['user_name'] or '',
                    row['user_email'] or '',
                    row['employee_id'] or '',
                ]).lower()
                if search not in haystack:
                    continue

            rows.append(row)

        return Response(rows)

    def perform_create(self, serializer):
        profile = serializer.save(tenant=self.get_tenant())
        if profile.primary_specialization_domain and profile.specialization != profile.primary_specialization_domain.name:
            profile.specialization = profile.primary_specialization_domain.name
            profile.save(update_fields=['specialization'])
        generate_faculty_eligible_subjects(profile)

    def perform_update(self, serializer):
        profile = serializer.save()
        if profile.primary_specialization_domain and profile.specialization != profile.primary_specialization_domain.name:
            profile.specialization = profile.primary_specialization_domain.name
            profile.save(update_fields=['specialization'])
        generate_faculty_eligible_subjects(profile)

    @action(detail=False, methods=['POST'], url_path='bulk-upload')
    def bulk_upload(self, request):
        """CSV bulk import for faculty.
        Required: email, employee_id, department_code, designation
        Optional: specialization, max_weekly_hours, joined_date, qualifications
        (User must already exist in the system)
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        tenant = self.get_tenant()
        csv_data = StringIO(file.read().decode('utf-8'))
        reader = csv.DictReader(csv_data)
        created, skipped, errors = 0, 0, []

        for row_num, row in enumerate(reader, start=2):
            email = row.get('email', '').strip().lower()
            emp_id = row.get('employee_id', '').strip()
            dept_code = row.get('department_code', '').strip().upper()
            _desig_map = {
                'professor': 'professor',
                'associate professor': 'associate_professor',
                'associate_professor': 'associate_professor',
                'assistant professor': 'assistant_professor',
                'assistant_professor': 'assistant_professor',
                'assistant': 'assistant_professor',
                'lecturer': 'lecturer',
                'visiting faculty': 'visiting',
                'visiting': 'visiting',
                'adjunct faculty': 'adjunct',
                'adjunct': 'adjunct',
                'head of department': 'hod',
                'hod': 'hod',
                'dean': 'dean',
                'other': 'other',
            }
            _raw_desig = row.get('designation', 'assistant_professor').strip().lower()
            designation = _desig_map.get(_raw_desig, 'assistant_professor')

            if not all([email, emp_id, dept_code]):
                errors.append(f'Row {row_num}: Missing email, employee_id, or department_code.')
                continue

            try:
                user = CustomUser.objects.get(email=email, tenant=tenant)
            except CustomUser.DoesNotExist:
                errors.append(f'Row {row_num}: User {email} not found in this tenant.')
                continue

            try:
                dept = Department.objects.get(tenant=tenant, code=dept_code)
            except Department.DoesNotExist:
                errors.append(f'Row {row_num}: Department "{dept_code}" not found.')
                continue

            if FacultyProfile.objects.filter(tenant=tenant, employee_id=emp_id).exists():
                skipped += 1
                errors.append(f'Row {row_num}: Faculty {emp_id} already exists (skipped).')
                continue

            try:
                FacultyProfile.objects.create(
                    tenant=tenant, user=user, department=dept,
                    employee_id=emp_id, designation=designation,
                    specialization=row.get('specialization', '').strip(),
                    max_weekly_hours=int(row.get('max_weekly_hours', 18)),
                    qualifications=row.get('qualifications', '').strip(),
                )
                created += 1
            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')

        return Response({'message': f'Done: {created} created, {skipped} skipped.', 'errors': errors})

    @action(detail=False, methods=['GET'], url_path='search-hod')
    def search_hod(self, request):
        """Lightweight endpoint for HOD combobox, searching faculty users."""
        q = request.query_params.get('q', '').strip()
        qs = self.get_queryset()
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(user__email__icontains=q))
        data = [{'id': f.user_id, 'name': f.user.full_name, 'email': f.user.email} for f in qs[:20]]
        return Response(data)

    @action(detail=False, methods=['GET', 'PATCH'], url_path='my-professional-profile')
    def my_professional_profile(self, request):
        """
        Faculty self-service endpoint.
        GET  — returns the logged-in faculty's professional profile fields.
        PATCH — updates the logged-in faculty's professional profile fields.
        Only faculty can use this endpoint.
        """
        user = request.user
        role_name = getattr(getattr(user, 'role', None), 'name', '')
        if role_name != 'faculty':
            return Response({'error': 'Only faculty members can access this endpoint.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            fp = user.faculty_profile
        except FacultyProfile.DoesNotExist:
            return Response({'error': 'No faculty profile found for your account.'}, status=status.HTTP_404_NOT_FOUND)

        department_ids = [fp.department_id] if fp.department_id else []
        excluded_faculty_domain_names = {
            'mathematics',
            'chemistry',
            'physics',
            'calculus & analysis',
            'linear algebra',
            'differential equations',
            'statistics & probability',
            'statistics & numerical methods',
            'life & environmental sciences',
            'electronics & electrical engineering',
            'electronics devices & circuits',
            'electrical engineering fundamentals',
            'general',
            'open elective',
            'professional elective',
            'internship',
            'project work',
            'management',
            'entrepreneurship',
            'ethics & human values',
            'humanities & management',
            'language & communication',
            'experiential learning',
        }

        if request.method == 'GET':
            serializer = FacultyProfileSerializer(fp)
            completion = profile_completion_state(fp)
            base_course_qs = Course.objects.filter(
                tenant=user.tenant,
                department_id__in=department_ids,
                primary_domain_id__isnull=False,
            ).select_related('primary_domain', 'secondary_domain', 'semester__program')

            if not base_course_qs.exists() and fp.department_id:
                normalized_name = self._normalize_department_name(fp.department.name)
                if normalized_name:
                    dept_qs = Department.objects.filter(tenant=user.tenant)
                    if fp.department.campus_id:
                        dept_qs = dept_qs.filter(campus_id=fp.department.campus_id)
                    family_ids = [
                        dept.id for dept in dept_qs
                        if self._normalize_department_name(dept.name) == normalized_name
                    ]
                    if family_ids:
                        department_ids = family_ids
                        base_course_qs = Course.objects.filter(
                            tenant=user.tenant,
                            department_id__in=department_ids,
                            primary_domain_id__isnull=False,
                        ).select_related('primary_domain', 'secondary_domain', 'semester__program')

            course_rows = list(base_course_qs)
            specialization_rows = [
                course for course in course_rows
                if getattr(getattr(course, 'semester', None), 'semester_number', 0) >= 3
            ] or course_rows

            allowed_domain_ids = set()
            for course in specialization_rows:
                if course.primary_domain_id:
                    allowed_domain_ids.add(str(course.primary_domain_id))
                if course.secondary_domain_id:
                    allowed_domain_ids.add(str(course.secondary_domain_id))

            domain_qs = SubjectDomain.objects.filter(
                tenant=user.tenant,
                department_id__in=department_ids,
                is_active=True,
            ).order_by('name')
            domains = [
                item for item in SubjectDomainSerializer(domain_qs, many=True).data
                if str(item.get('id')) in allowed_domain_ids and (item.get('name') or '').strip().lower() not in excluded_faculty_domain_names
            ]
            primary_domain_ids = []
            secondary_domain_ids = []
            related_secondary_domains = {}
            for course in specialization_rows:
                key = str(course.primary_domain_id)
                if key not in primary_domain_ids:
                    primary_domain_ids.append(key)
                related_secondary_domains.setdefault(key, [])
                if course.secondary_domain_id:
                    secondary_id = str(course.secondary_domain_id)
                    if secondary_id not in secondary_domain_ids:
                        secondary_domain_ids.append(secondary_id)
                    if secondary_id not in related_secondary_domains[key]:
                        related_secondary_domains[key].append(secondary_id)
            primary_domains = [domain for domain in domains if str(domain['id']) in primary_domain_ids]
            if not primary_domains:
                primary_domains = [domain for domain in domains if str(domain['id']) in secondary_domain_ids]
            if not primary_domains:
                primary_domains = domains
            allowed_domain_ids = {str(item['id']) for item in domains}
            related_secondary_domains = {
                key: [secondary_id for secondary_id in values if secondary_id in allowed_domain_ids]
                for key, values in related_secondary_domains.items()
                if key in allowed_domain_ids
            }
            return Response({
                **serializer.data,
                'available_subject_domains': domains,
                'available_primary_subject_domains': primary_domains,
                'related_secondary_domains': related_secondary_domains,
                'profile_completion_message': 'Complete your profile to receive subject preferences.' if not completion['ready_for_preferences'] else 'Profile is ready for subject preference selection.',
            })

        # PATCH — partial update of professional fields only
        ALLOWED_FIELDS = {
            'specialization', 'qualifications', 'skills', 'years_of_experience',
            'certifications', 'research_interests', 'industry_experience', 'bio',
            'primary_specialization_domain', 'secondary_specialization_domain_ids',
        }
        data = {k: v for k, v in request.data.items() if k in ALLOWED_FIELDS}
        serializer = FacultyProfileSerializer(fp, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        profile = serializer.save()
        generate_faculty_eligible_subjects(profile)
        return Response(FacultyProfileSerializer(profile).data)

    @action(detail=True, methods=['GET'], url_path='professional-detail')
    def professional_detail(self, request, pk=None):
        """
        Admin / HOD read-only view of a single faculty's professional details.
        Access control:
        - tenant_admin / super_admin: any faculty in the tenant
        - it_admin (campus admin):    faculty within their assigned campus
        - hod:                        faculty within their assigned department
        """
        user = request.user
        role_name = getattr(getattr(user, 'role', None), 'name', '')

        ALLOWED_ROLES = {'super_admin', 'tenant_admin', 'it_admin', 'hod', 'academic_admin'}
        if role_name not in ALLOWED_ROLES:
            return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            fp = FacultyProfile.objects.select_related(
                'user', 'department__campus'
            ).get(pk=pk, tenant=self.get_tenant())
        except FacultyProfile.DoesNotExist:
            return Response({'error': 'Faculty profile not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Campus admin scope
        if role_name == 'it_admin' and user.campus_id:
            if not fp.department or str(fp.department.campus_id) != str(user.campus_id):
                return Response({'error': 'Access denied: faculty is not in your campus.'}, status=status.HTTP_403_FORBIDDEN)

        # HOD scope — department match
        if role_name == 'hod':
            hod_department = self._get_hod_department()
            if not hod_department or (str(fp.department_id) != str(hod_department.id) and fp.department != hod_department):
                return Response({'error': 'Access denied: faculty is not in your department.'}, status=status.HTTP_403_FORBIDDEN)

        serializer = FacultyProfessionalDetailSerializer(fp)
        return Response(serializer.data)


class FacultyAvailabilityViewSet(TenantAcademicViewSet):
    serializer_class = FacultyAvailabilitySerializer
    filterset_fields = ['faculty', 'day', 'is_available']
    ordering_fields = ['day', 'start_time']

    def get_queryset(self):
        return FacultyAvailability.objects.filter(tenant=self.get_tenant()).select_related('faculty__user')

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())


class FacultyPreferenceViewSet(TenantAcademicViewSet):
    serializer_class = FacultyPreferenceSerializer
    filterset_fields = [
        'faculty', 'faculty__department', 'faculty__department__campus',
        'faculty__department_id', 'faculty__department__campus_id'
    ]

    def get_queryset(self):
        qs = FacultyPreference.objects.filter(tenant=self.get_tenant()).select_related(
            'faculty__user', 'faculty__department'
        ).prefetch_related('preferred_courses', 'ranked_subjects__course')
        # Faculty can only see their own preference
        user = self.request.user
        role_name = getattr(getattr(user, 'role', None), 'name', '')
        hod_department = self._get_hod_department()
        if role_name == 'faculty':
            try:
                fp = user.faculty_profile
                qs = qs.filter(faculty=fp)
            except Exception:
                if hod_department:
                    qs = qs.filter(faculty__department=hod_department)
                else:
                    return qs.none()
        elif hod_department and not self._is_admin():
            qs = qs.filter(faculty__department=hod_department)
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        role_name = getattr(getattr(user, 'role', None), 'name', '')
        if role_name in ['tenant_admin', 'academic_admin', 'super_admin']:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Tenant admins cannot create faculty preferences. Faculty must submit their own.")
        faculty = serializer.validated_data.get('faculty')
        if faculty:
            completion = profile_completion_state(faculty)
            if not completion['ready_for_preferences']:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'detail': 'Complete specialization and years of experience before submitting preferences.'})
        serializer.save(tenant=self.get_tenant())

    def perform_update(self, serializer):
        user = self.request.user
        role_name = getattr(getattr(user, 'role', None), 'name', '')
        if role_name in ['tenant_admin', 'academic_admin', 'super_admin']:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Tenant admins cannot edit faculty preferences.")
        faculty = serializer.instance.faculty
        completion = profile_completion_state(faculty)
        if not completion['ready_for_preferences']:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'detail': 'Complete specialization and years of experience before submitting preferences.'})
        serializer.save()

    def _filtered_preference_courses(self, faculty):
        selected_domain_ids = set()
        if getattr(faculty, 'primary_specialization_domain_id', None):
            selected_domain_ids.add(faculty.primary_specialization_domain_id)
        selected_domain_ids.update(faculty.secondary_specialization_domains.values_list('id', flat=True))
        if not selected_domain_ids:
            return Course.objects.none()
        scope_ids = [faculty.department_id] if faculty.department_id else []
        if faculty.department_id:
            normalized_name = self._normalize_department_name(faculty.department.name)
            if normalized_name:
                dept_qs = Department.objects.filter(tenant=faculty.tenant)
                if faculty.department.campus_id:
                    dept_qs = dept_qs.filter(campus_id=faculty.department.campus_id)
                for dept in dept_qs:
                    if self._normalize_department_name(dept.name) == normalized_name and dept.id not in scope_ids:
                        scope_ids.append(dept.id)
        return Course.objects.filter(
            tenant=faculty.tenant,
            department_id__in=scope_ids,
            is_active=True,
        ).filter(
            Q(primary_domain_id__in=selected_domain_ids) |
            Q(secondary_domain_id__in=selected_domain_ids)
        ).select_related(
            'semester__program', 'primary_domain', 'secondary_domain'
        ).order_by(
            'semester__program__name',
            'semester__semester_number',
            'primary_domain__name',
            'code',
        )

    def _serialize_preference_courses(self, courses):
        return [
            {
                'id': str(course.id),
                'course': str(course.id),
                'course_code': course.code,
                'course_name': course.name,
                'program_name': getattr(getattr(course.semester, 'program', None), 'name', None),
                'semester_name': getattr(course.semester, 'name', None),
                'primary_domain_name': getattr(course.primary_domain, 'name', None),
                'secondary_domain_name': getattr(course.secondary_domain, 'name', None),
                'status_display': 'Available',
            }
            for course in courses
        ]

    @action(detail=False, methods=['GET', 'POST', 'PATCH'], url_path='my-preference')
    def my_preference(self, request):
        """
        Faculty endpoint: GET to fetch own preference, POST/PATCH to create/update it.
        """
        user = request.user
        try:
            fp = user.faculty_profile
        except Exception:
            return Response({'error': 'No faculty profile found for this user.'}, status=status.HTTP_404_NOT_FOUND)

        pref, created = FacultyPreference.objects.get_or_create(
            faculty=fp, tenant=user.tenant
        )
        completion = profile_completion_state(fp)
        available_courses = self._filtered_preference_courses(fp)
        available_subjects = self._serialize_preference_courses(available_courses)

        if request.method == 'GET':
            return Response({
                **FacultyPreferenceSerializer(pref).data,
                'profile_completion': completion,
                'available_preference_subjects': available_subjects,
                'approved_eligible_subjects': available_subjects,
                'can_submit_preferences': completion['ready_for_preferences'],
            })

        # POST / PATCH — upsert preference
        if not completion['ready_for_preferences']:
            return Response(
                {
                    'detail': 'Complete specialization and years of experience before submitting preferences.',
                    'profile_completion': completion,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = FacultyPreferenceSerializer(pref, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        pref.refresh_from_db()
        return Response({
            **FacultyPreferenceSerializer(pref).data,
            'profile_completion': completion,
            'available_preference_subjects': available_subjects,
            'approved_eligible_subjects': available_subjects,
            'can_submit_preferences': completion['ready_for_preferences'],
        })

    @action(detail=False, methods=['GET'], url_path='my-courses')
    def my_courses(self, request):
        """
        Returns only HOD-approved eligible subjects for the faculty.
        """
        user = request.user
        try:
            fp = user.faculty_profile
        except Exception:
            return Response([])
        completion = profile_completion_state(fp)
        if not completion['ready_for_preferences']:
            return Response([])
        pool = self._filtered_preference_courses(fp)
        return Response(self._serialize_preference_courses(pool))

    @action(detail=True, methods=['POST'], url_path='approve')
    def approve(self, request, pk=None):
        if not self._get_hod_department() and not self._is_admin():
            return Response({'detail': 'Only HOD or admin can review preferences.'}, status=status.HTTP_403_FORBIDDEN)
        pref = self.get_object()
        ranked_subjects = list(pref.ranked_subjects.select_related('course'))
        pref.status = 'hod_approved'
        pref.hod_review_note = request.data.get('hod_review_note', '')
        pref.hod_reviewed_by = request.user
        pref.hod_reviewed_at = timezone.now()
        pref.save(update_fields=['status', 'hod_review_note', 'hod_reviewed_by', 'hod_reviewed_at', 'updated_at'])
        for ranked in ranked_subjects:
            FacultySubjectAssignment.objects.update_or_create(
                tenant=self.get_tenant(),
                faculty=pref.faculty,
                course=ranked.course,
                defaults={
                    'assigned_by': request.user,
                    'status': 'finalized',
                    'notes': f'Approved from faculty preference rank {ranked.rank}.',
                }
            )
        return Response(FacultyPreferenceSerializer(pref).data)

    @action(detail=True, methods=['POST'], url_path='reject')
    def reject(self, request, pk=None):
        if not self._get_hod_department() and not self._is_admin():
            return Response({'detail': 'Only HOD or admin can review preferences.'}, status=status.HTTP_403_FORBIDDEN)
        pref = self.get_object()
        pref.status = 'hod_rejected'
        pref.hod_review_note = request.data.get('hod_review_note', '')
        pref.hod_reviewed_by = request.user
        pref.hod_reviewed_at = timezone.now()
        pref.save(update_fields=['status', 'hod_review_note', 'hod_reviewed_by', 'hod_reviewed_at', 'updated_at'])
        return Response(FacultyPreferenceSerializer(pref).data)


class FacultyEligibleSubjectViewSet(TenantAcademicViewSet):
    serializer_class = FacultyEligibleSubjectSerializer
    filterset_fields = ['faculty', 'course', 'status', 'source_type']
    search_fields = ['faculty__user__full_name', 'course__code', 'course__name']
    ordering_fields = ['faculty__user__full_name', 'course__code', 'status']

    def get_queryset(self):
        if not self._get_hod_department() and not self._is_admin():
            return FacultyEligibleSubject.objects.none()
        qs = FacultyEligibleSubject.objects.filter(tenant=self.get_tenant()).select_related(
            'faculty__user', 'faculty__primary_specialization_domain',
            'course__semester__program', 'course__primary_domain', 'course__secondary_domain'
        )
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            scope_ids = self._get_hod_department_scope_ids()
            qs = qs.filter(faculty__department_id__in=scope_ids, course__department_id__in=scope_ids)
        department_id = self.request.query_params.get('department_id')
        faculty_id = self.request.query_params.get('faculty_id')
        program_id = self.request.query_params.get('program_id')
        domain_id = self.request.query_params.get('domain_id')
        if department_id:
            qs = qs.filter(faculty__department_id=department_id, course__department_id=department_id)
        if faculty_id:
            qs = qs.filter(faculty_id=faculty_id)
        if program_id:
            qs = qs.filter(course__semester__program_id=program_id)
        if domain_id:
            qs = qs.filter(Q(course__primary_domain_id=domain_id) | Q(course__secondary_domain_id=domain_id))
        return qs

    def create(self, request, *args, **kwargs):
        if not self._get_hod_department() and not self._is_admin():
            return Response({'detail': 'Only HOD or admin can manage eligible subjects.'}, status=status.HTTP_403_FORBIDDEN)
        faculty_id = request.data.get('faculty')
        course_id = request.data.get('course')
        if not faculty_id or not course_id:
            return Response({'detail': 'faculty and course are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            faculty = FacultyProfile.objects.get(id=faculty_id, tenant=self.get_tenant())
            course = Course.objects.get(id=course_id, tenant=self.get_tenant())
        except (FacultyProfile.DoesNotExist, Course.DoesNotExist):
            return Response({'detail': 'Faculty or course not found.'}, status=status.HTTP_404_NOT_FOUND)

        scope_ids = set(str(item) for item in self._get_hod_department_scope_ids())
        if scope_ids and (str(faculty.department_id) not in scope_ids or str(course.department_id) not in scope_ids):
            return Response({'detail': 'Faculty and course must belong to your department.'}, status=status.HTTP_403_FORBIDDEN)

        obj, _ = FacultyEligibleSubject.objects.update_or_create(
            tenant=self.get_tenant(),
            faculty=faculty,
            course=course,
            defaults={
                'source_type': 'hod_added',
                'status': 'hod_added',
                'notes': request.data.get('notes', ''),
                'score': request.data.get('score') or None,
            }
        )
        return Response(self.get_serializer(obj).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['POST'], url_path='regenerate')
    def regenerate(self, request):
        if not self._get_hod_department() and not self._is_admin():
            return Response({'detail': 'Only HOD or admin can regenerate eligible subjects.'}, status=status.HTTP_403_FORBIDDEN)
        regenerate_department_eligibility(self.get_tenant(), self._get_hod_department_scope_ids())
        return Response({'detail': 'Eligible subject pool regenerated.'})

    @action(detail=True, methods=['POST'], url_path='approve')
    def approve(self, request, pk=None):
        obj = self.get_object()
        obj.status = 'hod_approved'
        obj.save(update_fields=['status', 'updated_at'])
        return Response(self.get_serializer(obj).data)

    @action(detail=True, methods=['POST'], url_path='reject')
    def reject(self, request, pk=None):
        obj = self.get_object()
        obj.status = 'hod_rejected'
        obj.save(update_fields=['status', 'updated_at'])
        return Response(self.get_serializer(obj).data)


class FacultySubjectAssignmentViewSet(TenantAcademicViewSet):
    serializer_class = FacultySubjectAssignmentSerializer
    filterset_fields = ['faculty', 'course', 'status']
    search_fields = ['faculty__user__full_name', 'course__code', 'course__name']
    ordering_fields = ['faculty__user__full_name', 'course__code']

    def get_queryset(self):
        if not self._get_hod_department() and not self._is_admin():
            return FacultySubjectAssignment.objects.none()
        qs = FacultySubjectAssignment.objects.filter(tenant=self.get_tenant()).select_related(
            'faculty__user', 'course__semester__program', 'assigned_by'
        )
        hod_department = self._get_hod_department()
        if hod_department and not self._is_admin():
            scope_ids = self._get_hod_department_scope_ids()
            qs = qs.filter(faculty__department_id__in=scope_ids, course__department_id__in=scope_ids)
        department_id = self.request.query_params.get('department_id')
        faculty_id = self.request.query_params.get('faculty_id')
        program_id = self.request.query_params.get('program_id')
        semester_number = self.request.query_params.get('semester_number')
        if department_id:
            qs = qs.filter(faculty__department_id=department_id, course__department_id=department_id)
        if faculty_id:
            qs = qs.filter(faculty_id=faculty_id)
        if program_id:
            qs = qs.filter(course__semester__program_id=program_id)
        if semester_number:
            qs = qs.filter(course__semester__semester_number=semester_number)
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant(), assigned_by=self.request.user)

    def create(self, request, *args, **kwargs):
        faculty_id = request.data.get('faculty')
        course_id = request.data.get('course')
        if faculty_id and course_id:
            allowed = FacultyEligibleSubject.objects.filter(
                tenant=self.get_tenant(),
                faculty_id=faculty_id,
                course_id=course_id,
                status__in=['hod_approved', 'hod_added'],
            ).exists()
            if not allowed:
                return Response({'detail': 'Final assignment is allowed only for HOD-approved eligible subjects.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().create(request, *args, **kwargs)



# ── Course Section (Assignment) ───────────────────────────────────────────────

class CourseSectionViewSet(TenantAcademicViewSet):
    serializer_class = CourseSectionSerializer
    filterset_fields = ['course', 'course_id', 'course__department', 'course__department_id', 'course__department__campus', 'course__department__campus_id', 'section', 'semester', 'faculty', 'is_active']
    search_fields = ['course__code', 'course__name', 'faculty__user__full_name']
    ordering_fields = ['course__code']

    def get_queryset(self):
        qs = CourseSection.objects.filter(tenant=self.get_tenant()).select_related(
            'course__department__campus', 'section__batch', 'semester', 'faculty__user', 'student_group'
        )

        user = self.request.user
        role_name = getattr(getattr(user, 'role', None), 'name', '')
        hod_department = self._get_hod_department()
        is_hod = bool(hod_department)
        is_admin = role_name in ['super_admin', 'tenant_admin', 'campus_admin', 'it_admin', 'academic_admin']

        campus_id = self.request.query_params.get('campus_id')
        department_id = self.request.query_params.get('department_id')
        course_id = self.request.query_params.get('course_id')
        section_id = self.request.query_params.get('section_id') or self.request.query_params.get('section')
        faculty_id = self.request.query_params.get('faculty_id') or self.request.query_params.get('faculty')

        # Handle explicit faculty filter (the "View More" case)
        if faculty_id:
            try:
                import uuid
                uuid.UUID(str(faculty_id))
            except (ValueError, TypeError):
                return qs.none()
            
            # Permission check: only admins and HODs can see other faculty's courses
            if not is_admin and not is_hod:
                my_profile_id = str(user.faculty_profile.id) if hasattr(user, 'faculty_profile') else None
                if str(faculty_id) != my_profile_id:
                    return qs.none()
            
            qs = qs.filter(faculty_id=faculty_id)
        
        # If no specific faculty filter, but role is faculty and NOT admin/HOD, default to own courses
        elif role_name == 'faculty' and not is_admin and not is_hod:
            if hasattr(user, 'faculty_profile'):
                qs = qs.filter(faculty=user.faculty_profile)
            else:
                return qs.none()
        elif is_hod and not is_admin:
            qs = qs.filter(course__department=hod_department)

        if campus_id:
            qs = qs.filter(course__department__campus_id=campus_id)
        if department_id:
            qs = qs.filter(course__department_id=department_id)
        if course_id:
            qs = qs.filter(course_id=course_id)
        if section_id:
            qs = qs.filter(section_id=section_id)
            
        return qs

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())

    @action(detail=False, methods=['GET'], url_path='by-faculty')
    def by_faculty(self, request):
        """Returns course assignments grouped readable by faculty."""
        faculty_id = request.query_params.get('faculty_id')
        qs = self.get_queryset()
        if faculty_id:
            qs = qs.filter(faculty_id=faculty_id)
        return Response(CourseSectionSerializer(qs, many=True).data)


# ── Dept Room Preference ──────────────────────────────────────────────────────

class DepartmentRoomPreferenceViewSet(TenantAcademicViewSet):
    serializer_class = DepartmentRoomPreferenceSerializer
    filterset_fields = ['department', 'course_type']

    def get_queryset(self):
        return DepartmentRoomPreference.objects.filter(tenant=self.get_tenant()).select_related('department', 'room')

    def perform_create(self, serializer):
        serializer.save(tenant=self.get_tenant())
